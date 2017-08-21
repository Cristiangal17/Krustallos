from openpyxl import Workbook #creats writable Workbook
from xlrd import * # Reads .xls
from xlutils.copy import copy
from collections import OrderedDict
import os


## Try not to use stored data

def columnNames( fileDir , row, sheetNum):
    '''This returns the column names of the file. You ave to specify where the column names are.'''
    adict = OrderedDict()
    with open_workbook(fileDir) as mainOpen:
        sheet = mainOpen.sheet_by_index(sheetNum)
        for c in range(sheet.ncols):
            adict[sheet.cell_value(row,c)] = c
        return adict

def copyTemplate(fileDir,tempName, nameOfOutput):
    current = os.getcwd()
    shutil.copy(fileDir, current)
    copyDir = os.path.join(current,tempName)
    os.rename(copyDir, nameOfOutput)
    copyDir = os.path.join(current,nameOfOutput)
    return copyDir


def invEnumerated(enumDict):
    adict = OrderedDict()
    for k,v in enumDict.items():
        adict[v] = k
    return adict


def indexedMap(tempNamesDict, mainFileDict, toMap):
    indexMap = []
    for i in toMap:
        #Can do a try and except then raise an error with a clearer message if needed
        tempIndex = tempNamesDict[i[0]]
        mainIndex = mainFileDict[i[1]]
        indexMap.append((tempIndex,mainIndex))
    return indexMap


if __name__ == '__main__':

    #This is the info that i will need from the use
    #Have a function that subtracts 1 from the user input for proper indexing sheet number and row number
    mainFileName = "B&Bitems169-263Filled.xls"
    fileLocal = os.path.join(os.getcwd(),"MainFiles", mainFileName)
    
    fileSheet = 3
    fileColNameRow = 9
    dataRowStart = 10
    #(TemplateMap,InfoMap) Here is where you tell the program what to map
    ##THE USER INTERFACE SHOULD GIVE ME SOMETHING LIKE THIS
    #("SKU","Item Name")
    toMap = [("Title","Amazon title"),("Manufacturer", "Vendor Name"), ("Model Number", "Item Name"),("Product Type", "Amazon Product Type"),("Item Type Keyword", "Amazon Item KeyWord"), ("Brand Name", "Vendor Name"),("Product Description","Website Item Description"), ("Standard Price", "Regular Price"), ("Quantity", "Qty 1"), ("Currency", "Currency"), ("Diameter","Diameter"),
             ("Item Length", "Item Length"),("Item Width","Item Width"),("Item Height","Item Height"),("Item Dimensions Unit Of Measure", "Dimensions Unit Of Measure")]
    toMap2 = [("Product Name","Amazon title"), ("Product Description","Website Item Description")]


    # Here is the information for the template
    templateName = "FlatFileJewelry.xls"
    templateName2 = "FlatFileHome.xlsm"
    template = os.path.join(os.getcwd(),"Templates", templateName2) #Change
    templateSheet = 3
    templateColNameRow = 1

    tempNames = columnNames(template, templateColNameRow,templateSheet)
    mainFileNames = columnNames(fileLocal, fileColNameRow, fileSheet)
    
    #Finds the columns where the info should be inserted
    indexMap = indexedMap(tempNames, mainFileNames, toMap2) #change

    #creates a new workbook
    wBook = Workbook()
    wSheet = wBook.active


    
    with open_workbook(fileLocal) as mainOpen:
        sheet = mainOpen.sheet_by_index(fileSheet)
        count = 1

        #names the columns of the workbook
        for k in tempNames:
            currentCell = wSheet.cell(row = 1, column = count)
            currentCell.value = k
            count += 1

        #maps the information from the main file to the workbook
        for i in indexMap:
            temp = i[0]
            main = i[1]
            for r in range(dataRowStart,sheet.nrows):
                currentCell = wSheet.cell(row = r - dataRowStart + 2, column = temp +1)
                currentCell.value = sheet.cell_value(r,main)

    #saves the workbook
    wBook.save("Test2.xls") #change
    print("SAVED!")
            

