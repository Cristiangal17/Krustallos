'''
Created on Jun 18, 2015

@author: cristiangalindo
'''
from openpyxl import load_workbook, Workbook
from xlrd import open_workbook



fileLocal = "/Users/cristiangalindo/Documents/workspace/Krustallos/Export.xls"
workbook = open_workbook(fileLocal)
sheet = workbook.sheet_by_index(0)
print(sheet.nrows) #Tells you how many rows
print(sheet.ncols) #Tells you how many columns
print(sheet.cell_value(0,0))

data = [[sheet.cell_value(r,c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
colNames = data[0]
adict = {}
rows = sheet.nrows
columns = sheet.ncols

wb = Workbook()
ws = wb.active

for i in range(len(data[0])):
    adict[i] = data[0][i]

##for k,v in adict.items():
##    print(k,v)
##print(adict[0])

#print(ws.columns)


for i in range(1,sheet.ncols):
    currentCell = ws.cell(row = 1, column = i)
    currentCell.value = adict[i]
wb.save('Test.xls')




##dataRowStart = 10
##with open_workbook(fileLocal) as mainOpen:
##    
##    sheet = mainOpen.sheet_by_index(fileSheet)
##    openTemp = open_workbook(dirOfCopy)
##    for i in indexMap:
##        temp = i[0]
##        main = i[1]
##        
##        for r in range(dataRowStart,sheet.nrows):
##            print(sheet.cell_value(r,main))
##        print("NEXT SECTION")
##


