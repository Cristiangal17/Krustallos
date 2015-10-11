from openpyxl import Workbook #creats writable Workbook
from xlrd import * # Reads .xls
from xlutils.copy import copy
from collections import OrderedDict, defaultdict
import shutil
import os
from PopulateAmazonTNoS import columnNames, invEnumerated

'''The user is allowed to specify what type of info he is passing.The user should be able to
use these two options:
-Inventory
-Sold
For the inventory option, the user is passing the file will all the inventory.
(I need to make sure all the numbers are correct. This options gives me how many are avaliable).

For the sold the user will pass on a template of the items that were sold.
(I need to add the information to the sold column and calculate the new ammount avaliable).'''



def inventoryDict(sheetInfo, dictionary, items,start):
    alist = [dictionary[items[0]],dictionary[items[1]]]
    adict = OrderedDict()
    cols = sheetInfo.ncols
    rows = sheetInfo.nrows
    for r in range(start,rows):
        currentRowItem = sheetInfo.cell_value(r , alist[0])
        currentRowAmm = sheetInfo.cell_value(r , alist[1])
        adict[currentRowItem] = currentRowAmm
    return adict
            
                
    

if __name__ == '__main__':


    amazonInfoFileName = "Amazon.xls"
    amazonInfoType = "Inventory"
    #amazonInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", amazonInfoFileName)
    amazonInfoLocal = os.path.join(os.getcwd(), "ToUpdate", amazonInfoFileName)
    amazonInfoSheet = 1
    aRowNameStart = 1
    aInfoStart = 2

    ebayInfoFileName = "Ebay.xls"
    ebayInfoType = "Inventory"
    #ebayInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", ebayInfoFileName)
    ebayInfoLocal = os.path.join(os.getcwd(), "ToUpdate", ebayInfoFileName)

    ebayInfoSheet = 1
    eRowNameStart = 1
    eInfoStart = 2

    websiteInfoFileName = "Website.xls"
    websiteInfoType = "Inventory"
    #websiteInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", websiteInfoFileName)
    websiteInfoLocal = os.path.join(os.getcwd(), "ToUpdate", websiteInfoFileName)
    websiteInfoSheet = 1
    wRowNameStart = 1
    wInfoStart = 2

    posInfoFile = "POS.xls"
    posInfoType = "Inventory"
    #posInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", posInfoFile)
    posInfoLocal = os.path.join(os.getcwd(), "ToUpdate", posInfoFile)

    posInfoSheet = 1
    pRowNameStart = 1
    pInfoStart = 2

    searsInfoFileName = "Sears.xls"
    #searsInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", searsInfoFileName)
    searsInfoLocal = os.path.join(os.getcwd(), "ToUpdate", searsInfoFileName)
    searsInfoSheet = 3
    sRowNameStart = 1
    sInfoStart = 2
    
    inventoryFileName = "AllInventory.xls"
    inventoryInfoLocal = os.path.join(os.getcwd(), "MainFiles", "InventoryFile", inventoryFileName)
    invSheetNum = 4
    invRowNameStart = 10
    invInfoStart = 11



    
    #Here is where i change the indexes to match
    amazonInfoSheet = amazonInfoSheet - 1
    aRowNameStart = aRowNameStart - 1
    aInfoStart = aInfoStart - 1

    ebayInfoSheet = ebayInfoSheet - 1
    eRowNameStart = eRowNameStart - 1
    eInfoStart = eInfoStart - 1
    
    websiteInfoSheet = websiteInfoSheet - 1
    wRowNameStart = wRowNameStart - 1
    wInfoStart = wInfoStart - 1

    posInfoSheet = posInfoSheet - 1
    pRowNameStart = pRowNameStart - 1
    pInfoStart = pInfoStart - 1
    
    searsInfoSheet = searsInfoSheet - 1
    sRowNameStart = sRowNameStart - 1
    sInfoStart = sInfoStart - 1

    invSheetNum = invSheetNum - 1
    invRowNameStart = invRowNameStart - 1
    invInfoStart = invInfoStart - 1
    #end of index change



    aInfoCols = ["sku","quantity"]
    amazonColNames = columnNames(amazonInfoLocal, aRowNameStart, amazonInfoSheet)
    amazonInfoOpen = open_workbook(amazonInfoLocal)
    amazonSheet = amazonInfoOpen.sheet_by_index(amazonInfoSheet)
    amaRows = amazonSheet.nrows
    amaCols = amazonSheet.ncols
    amazonInv = inventoryDict(amazonSheet, amazonColNames, aInfoCols, aInfoStart)
    print("Amazon INV")
    print(amazonInv)
    print()


    eInfoCols = ["CustomLabel", "Quantity"]
    ebayColNames = columnNames(ebayInfoLocal, eRowNameStart, ebayInfoSheet)
    ebayInfoOpen = open_workbook(ebayInfoLocal)
    ebaySheet = ebayInfoOpen.sheet_by_index(ebayInfoSheet)
    ebayRows = ebaySheet.nrows
    ebayCols = ebaySheet.ncols
    ebayInv = inventoryDict(ebaySheet, ebayColNames, eInfoCols, eInfoStart)
    print("Ebay INV")
    print(ebayInv)
    print()

    

    webInfoCols = ["Product ID", "Stock"]
    webColNames = columnNames(websiteInfoLocal, wRowNameStart, websiteInfoSheet)
    websiteInfoOpen = open_workbook(websiteInfoLocal)
    websiteSheet = websiteInfoOpen.sheet_by_index(websiteInfoSheet)
    webRows = websiteSheet.nrows
    webCols = websiteSheet.ncols
    webInv = inventoryDict(websiteSheet, webColNames, webInfoCols, wInfoStart)
    print("Website INV")
    print(webInv)  
    print()

    posInfoCols = ["Item Name", "Qty 1"]
    posColNames = columnNames(posInfoLocal, pRowNameStart,posInfoSheet)
    posInfoOpen = open_workbook(posInfoLocal)
    posSysSheet = posInfoOpen.sheet_by_index(posInfoSheet)
    posRows = posSysSheet.nrows
    posCols = posSysSheet.ncols
    posInv = inventoryDict(posSysSheet, posColNames, posInfoCols, pInfoStart)
    print("POS Inv")
    print(posInv)
    print()
    
    
    
    searsInfoCols = ["Item Id", "Existing Available Quantity"]
    searsColNames = columnNames(searsInfoLocal, sRowNameStart, searsInfoSheet)
    searsInfoOpen = open_workbook(searsInfoLocal)
    searsOGSheet = searsInfoOpen.sheet_by_index(searsInfoSheet)
    searsRows = searsOGSheet.nrows
    searsCols = searsOGSheet.ncols
    searsInv = inventoryDict(searsOGSheet, searsColNames, searsInfoCols, sInfoStart)
    
    
    
    
    newWorkbook = Workbook()
    wSheet = newWorkbook.active

################################################################################
####HERE IS WHERE IT GETS ALL THE INVENTORY AND CREATES A NEW TEMPLATE####


    incorrectDict = defaultdict()
    
    with open_workbook(inventoryInfoLocal) as openInventory:
        inventorySheet = openInventory.sheet_by_index(invSheetNum)
        rows = inventorySheet.nrows
        cols = inventorySheet.ncols
        
        columnNames = columnNames(inventoryInfoLocal, invRowNameStart, invSheetNum)
        #print(columnNames)
        revColNames = invEnumerated(columnNames)

        ### MAKE A DICT OF NAMES THAT DON'T EXIST IN THE STORE FILES BUT WAS PUT ON THE MAIN FILE
        
        for k,v in columnNames.items():
            
            currentCell = wSheet.cell(row = 1, column = v + 1)
            currentCell.value = k
            
        oldn = columnNames["Old Product ID"]
        webn = columnNames["Website Item Name"]
        posn = columnNames["POS Item Name"]
        ebayn = columnNames["Ebay Custom Label"]
        aman = columnNames["Amazon Sku"]
        sean = columnNames["Sears Name"]
        start = columnNames["Starting Quantity"]
        short = columnNames["Short Description"]
        sstore = columnNames["Sold in Store"]
        sama = columnNames["Sold in Amazon"]
        sebay = columnNames["Sold in Ebay"]
        swebsite = columnNames["Sold in Website"]
        ssears = columnNames["Sold in Sears"]
        tsold = columnNames["Total Sold"]
        qty = columnNames["Qty 1"]
        
        
        
        for r in range(invInfoStart,rows):
            
            
            itemPos = ""
            itemAmazon = ""
            itemWebsite = ""
            itemEbay = ""
            itemSears = ""
            
            
            startAmount = 0
            avaliablePos = 0
            avaliableWeb = 0
            avaliableAma = 0
            avaliableEbay = 0
            avaliableSears = 0

            soldPos = 0
            soldAma = 0
            soldWeb = 0
            soldEbay = 0
            soldSears = 0
            totalSold = 0
            finalAmount = 0

            
            for c in range(cols):
                #print(r,c)
                currentCell = wSheet.cell(row = r - invInfoStart + 2 ,column = c+1)
                if c == columnNames["POS Item Name"]:
                    try:
                        currentCell.value = inventorySheet.cell_value(r,c)
                        avaliablePos = int(posInv[inventorySheet.cell_value(r,c)])
                        itemPos = inventorySheet.cell_value(r,c)
                    except KeyError:
                        if 'pos' in incorrectDict.keys():
                            incorrectDict['pos'].append(inventorySheet.cell_value(r,c))
                        else:
                            incorrectDict['pos'] = [inventorySheet.cell_value(r,c)]
                            
                if c == columnNames["Old Product ID"] or c == columnNames["Price"] or c == columnNames["Type"]:
                    currentCell.value = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Amazon Sku"]:
                    try:
                        currentCell.value = inventorySheet.cell_value(r,c)
                        avaliableAma = int(amazonInv[inventorySheet.cell_value(r,c)])
                        itemAmazon = inventorySheet.cell_value(r,c)
                    except KeyError:
                        if 'amazon' in incorrectDict.keys():
                            incorrectDict['amazon'].append(inventorySheet.cell_value(r,c))
                        else:
                            incorrectDict['amazon'] = [inventorySheet.cell_value(r,c)]
                        continue
                    
                if c == columnNames["Website Item Name"]:
                    try:
                        currentCell.value = inventorySheet.cell_value(r,c)
                        avaliableWeb = int(webInv[inventorySheet.cell_value(r,c)])
                        itemWebsite = inventorySheet.cell_value(r,c)
                        
                    except KeyError:
                        if 'website' in incorrectDict.keys():
                            incorrectDict['website'].append(inventorySheet.cell_value(r,c))
                        else:
                            incorrectDict['website'] = [inventorySheet.cell_value(r,c)]
                    
                if c == columnNames["Ebay Custom Label"]:
                    try:
                        currentCell.value = inventorySheet.cell_value(r,c)
                        avaliableEbay = int(ebayInv[inventorySheet.cell_value(r,c)])
                        itemEbay = inventorySheet.cell_value(r,c)
                        
                    except KeyError:
                        if 'ebay' in incorrectDict.keys():
                            incorrectDict['ebay'].append(inventorySheet.cell_value(r,c))
                        else:
                            incorrectDict['ebay'] = [inventorySheet.cell_value(r,c)]
                
                if c == columnNames["Sears Name"]:
                    try:
                        currentCell.value = inventorySheet.cell_value(r,c)
                        avaliableSears = int(searsInv[inventorySheet.cell_value(r,c)])
                        itemSears = inventorySheet.cell_value(r,c)
                    except KeyError:
                        if 'sears' in incorrectDict.keys():
                            incorrectDict['sears'].append(inventorySheet.cell_value(r,c))
                        else:
                            incorrectDict['sears'] = [inventorySheet.cell_value(r,c)]
                    
                if c == columnNames["Short Description"]:
                    currentCell.value = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Starting Quantity"]:
                    startAmount = inventorySheet.cell_value(r ,c)
                    currentCell.value = startAmount

                if c == columnNames["Sold in Store"]:
                    soldPos = inventorySheet.cell_value(r ,c)
                    if itemPos != "":
                        soldPos = soldPos + (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliablePos)
                        if (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliablePos) != 0 : print("Sold In Store", itemPos)
                        currentCell.value = soldPos
                    if itemPos == "":
                        soldPos = 0
                        currentCell.value = soldPos
                            
                if c == columnNames["Sold in Amazon"]:
                    soldAma = inventorySheet.cell_value(r ,c)
                    if itemAmazon != "":
                        soldAma = soldAma + (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableAma)
                        if (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableAma) != 0 : print("Sold In Amazon", itemAmazon)
                        
                        currentCell.value = soldAma
                    if itemAmazon == "":
                        soldAma = 0
                        currentCell.value = soldAma

                        
                        
                if c == columnNames["Sold in Ebay"]:
                    soldEbay = inventorySheet.cell_value(r ,c)
                    if itemEbay != "":
                        soldEbay = soldEbay + (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableEbay)
                        if (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableEbay) != 0 : print("Sold In Ebay", itemEbay)
                        currentCell.value = soldEbay
                    if itemEbay == "":
                        soldEbay = 0
                        currentCell.value = soldEbay
                    
                if c == columnNames["Sold in Website"]:
                    soldWeb = inventorySheet.cell_value(r ,c)
                    if itemWebsite != "":
                        soldWeb = soldWeb + (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableWeb)
                        if (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableWeb) != 0 : print("Sold In Website", itemWebsite)

                        currentCell.value = soldWeb
                    if itemWebsite == "":
                        soldWeb = 0
                        currentCell.value = soldWeb

                            
                if c == columnNames["Sold in Sears"]:
                    soldSears = inventorySheet.cell_value(r ,c)
                    if itemSears != "":
                        soldSears = soldSears + (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableSears)
                        if (inventorySheet.cell_value(r ,columnNames["Qty 1"]) - avaliableSears) != 0 : print("Sold In Sears", itemSears)
                        currentCell.value = soldSears
                    if itemSears == "":
                        soldSears = 0
                        currentCell.value = soldSears
                        
                
                if c == columnNames["Total Sold"]: 
                    totalSold = soldPos + soldAma + soldEbay + soldWeb + soldSears
                    currentCell.value = totalSold 

                        
                        
                if c == columnNames["Qty 1"]:
                    currentCell.value = startAmount - totalSold

                


    ## saves the error log
                    
    errorLog = "MasterFileNamesErrorLog.txt"
    with open(errorLog, 'w') as errors:
        for k,v in incorrectDict.items():
            errors.write("Wrong Id in the  store " + k + "\n")
            for item in v:
                errors.write(item + "\n")
                
    os.rename(os.path.join(os.getcwd(),errorLog),os.path.join(os.getcwd(), "errors", errorLog))
    newWorkbook.save("Inventory.xls")
    print("Saved!")

################################################################################
### HERE IS WHERE IT SHOULD CREATE THE TEMPLATES TO REUPLOAD###
    
               
    amaInvUpdateWb = Workbook()
    amaSheet = amaInvUpdateWb.active

    ebayInvUpdateWb = Workbook()
    ebaySheet = ebayInvUpdateWb.active

    webInvUpdateWb = Workbook()
    webSheet = webInvUpdateWb.active

    posInvUpdateWb = Workbook()
    posSheet = posInvUpdateWb.active

    searsInvUpdateWb = Workbook()
    searsSheet = searsInvUpdateWb.active
    

################################################################################
### HERE IS WHERE IT SHOULD repopulate THE TEMPLATES TO REUPLOAD###

    
    newPosInv = OrderedDict()
    newAmaInv = OrderedDict()
    newEbayInv = OrderedDict()
    newWebInv = OrderedDict()
    newSearsInv = OrderedDict()
    
    
    productNotOnMaster = OrderedDict()
        
    newInvFileLocal = os.path.join(os.getcwd(),"Inventory.xls")
    newInvSheetNum = 0
    newInvStart = 1
    

    
    #This part goes through the newly populated template and creates dictionaries that
    #Will be used to create the new inventory for the new templates
    with open_workbook(newInvFileLocal) as openInventory:
        inventorySheet = openInventory.sheet_by_index(newInvSheetNum)
        rows = inventorySheet.nrows
        cols = inventorySheet.ncols
        for r in range(newInvStart,rows):
            posName = ""
            amazonName = ""
            ebayName = ""
            websiteName = ""
            searsName = ""
            
            for c in range(cols):
                currentCell = wSheet.cell(row = r + 1 ,column = c+1)
                #print(inventorySheet.cell_value(r,c))


                if c == columnNames["POS Item Name"] :
                    posName = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Amazon Sku"]:
                    amazonName = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Ebay Custom Label"]:
                    ebayName = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Website Item Name"]:
                    websiteName = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Sears Name"]:
                    searsName = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Qty 1"]:
                    finalAmount = inventorySheet.cell_value(r ,c)
                    newPosInv[posName] = finalAmount
                    newAmaInv[amazonName] = finalAmount
                    newEbayInv[ebayName] = finalAmount
                    newWebInv[websiteName] = finalAmount
                    newSearsInv[searsName] = finalAmount


##################################################################################
###Repupulates the new amazon file with the new inventory###

    ###MAKE A DICT OF NAMES THAT EXIST ON THE FILES BUT NOT ON THE MAIN FILE
                    
    amaCount = 1
    amaReuploadCols = OrderedDict()
    for k in amazonColNames.keys():
        if k.lower() != "asin":
            currCell = amaSheet.cell(row = 1, column = amaCount)
            currCell.value = k.lower()
            amaReuploadCols[k] = amaCount
            amaCount += 1


    prow = 1
    for r in range(1,amaRows):
        quantity = 0
        for c in range(amaCols):
            cellValue = amazonSheet.cell_value(r,c)
            try:
                if c == amazonColNames['sku']:
                    currCell = amaSheet.cell(row = prow + 1, column = amaReuploadCols['sku']) 
                    quantity = newAmaInv[cellValue]
                    currCell.value = cellValue
                if c == amazonColNames['price']:
                    currCell = amaSheet.cell(row = prow + 1, column = amaReuploadCols['price'])
                    currCell.value = cellValue
                if c == amazonColNames['quantity']:
                    currCell = amaSheet.cell(row = prow + 1, column = amaReuploadCols['quantity'])
                    currCell.value = quantity
            except KeyError:
                prow -= 1
                ## Create files that will tell us what product ids are on the file that are not on the main file
                #print("Caught " , cellValue)
                if 'amazon' in productNotOnMaster.keys():
                    productNotOnMaster['amazon'].append(cellValue)
                    
                else:
                    productNotOnMaster['amazon'] = [cellValue]
                break
                
        prow += 1
                
##################################################################################
###Repupulates the new ebay file with the new inventory##l
    for k,v in ebayColNames.items():
        currCell = ebaySheet.cell(row = 1, column = v + 1)
        currCell.value = k




##################################################################################
###Repupulates the new website file with the new inventory###
    for k,v in webColNames.items():
        currCell = webSheet.cell(row = 1, column = v + 1)
        currCell.value = k

    prow = 1 
    for r in range(1,webRows):
        quantity = 0
        for c in range(webCols):
            cellValue = websiteSheet.cell_value(r,c)
            currCell = webSheet.cell(row = prow + 1, column = c + 1)
            try:
                if c == webColNames['Product ID']:
                    quantity = newWebInv[cellValue]
                    currCell.value = cellValue
                    
                if c == webColNames['Stock']:
                    currCell.value = quantity
                    
                if c != webColNames['Stock'] and c != webColNames['Product ID']:
                    currCell.value = cellValue
            except KeyError:
                prow -= 1
                #print(cellValue)
                if 'website' in productNotOnMaster.keys():
                    productNotOnMaster['website'].append(cellValue)
                else:
                    productNotOnMaster['website'] = [cellValue]
                
        prow += 1
                    
                    

##################################################################################
###Repupulates the new pos file with the new inventory###
        
    for k,v in posColNames.items():
        currCell = posSheet.cell(row = 1, column = v + 1)
        currCell.value = k
    
    prow = 1
    for r in range(1,posRows):
        try:
            itemName = posSysSheet.cell_value(r,posColNames['Item Name'])
            quantity = newPosInv[itemName]
            for c in range(posCols):
                cellValue = posSysSheet.cell_value(r,c)
                currCell = posSheet.cell(row = prow + 1, column = c + 1)

                ##if a product is on the pos sys but not in the master file it will
                ##write the first column number, change that

                if c == posColNames['Qty 1']:
                    currCell.value = quantity

                if c != posColNames['Qty 1']:
                    currCell.value = cellValue
        except KeyError:
            prow -= 1
            #print(itemName)
            if 'pos' in productNotOnMaster.keys():
                productNotOnMaster['pos'].append(itemName)
            else:
                productNotOnMaster['pos'] = [itemName]
                
        prow += 1
                
            #print(cellValue)
##################################################################################
###Repupulates the new SEARS file with the new inventory###        
    
    for k,v in searsColNames.items():
        currCell = searsSheet.cell(row = 1, column = v + 1)
        currCell.value = k
         
    prow = 1
    for r in range(1, searsRows):
        try:
            itemName = searsOGSheet.cell_value(r, searsColNames["Item Id"])
            quantity = newSearsInv[itemName]
            for c in range(searsCols):
                cellValue = searsOGSheet.cell_value(r,c)
                currCell = searsSheet.cell(row = prow + 1, column = c + 1)
                if c == searsColNames['Updated Available Quantity']:
                    currCell.value = quantity
                     
                if c != searsColNames['Updated Available Quantity']:
                    currCell.value = cellValue
             
             
        except KeyError:
            prow -= 1
            #print(itemName)
            if 'sears' in productNotOnMaster.keys():
                productNotOnMaster['sears'].append(itemName)
            else:
                productNotOnMaster['sears'] = [itemName]
                 
        prow += 1
     
     
                    


    missingLog = "missingInMaster.txt"

    with open(missingLog, 'w') as missing:
        for k,v in productNotOnMaster.items():
            missing.write("Missing Id in the  store " + k + "\n")
            for item in v:
                missing.write(item + "\n")
    os.rename(os.path.join(os.getcwd(),missingLog),os.path.join(os.getcwd(), "errors", missingLog))


##################################################################################
###HERE IS THE NAMING OF THE FILES###


    #print(productNotOnMaster)
    amaUpdateFileName = "AmazonInvUpdate.xls"
    ebayUpdateFileName = "EbayInvUpdate.xls"
    webUpdateFileName = "WebInvUpdate.xls"
    posUpdateFileName = "POSInvUpdate.xls"
    searsUpdateFileName = "SearsInvUpdate.xls"
    
    amaInvUpdateWb.save(amaUpdateFileName)
    ebayInvUpdateWb.save(ebayUpdateFileName)
    webInvUpdateWb.save(webUpdateFileName)
    posInvUpdateWb.save(posUpdateFileName)
    searsInvUpdateWb.save(searsUpdateFileName)


##################################################################################
###HERE IS WHERE THE FILES ARE MOVED###
    os.rename(os.path.join(os.getcwd(),amaUpdateFileName),os.path.join(os.getcwd(), "Updates", amaUpdateFileName))
    os.rename(os.path.join(os.getcwd(),ebayUpdateFileName),os.path.join(os.getcwd(), "Updates", ebayUpdateFileName))
    os.rename(os.path.join(os.getcwd(),webUpdateFileName),os.path.join(os.getcwd(), "Updates", webUpdateFileName))
    os.rename(os.path.join(os.getcwd(),posUpdateFileName),os.path.join(os.getcwd(), "Updates", posUpdateFileName))
    os.rename(os.path.join(os.getcwd(),searsUpdateFileName),os.path.join(os.getcwd(), "Updates", searsUpdateFileName))

    

    print("Saved updates!")

    

