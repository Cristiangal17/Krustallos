from openpyxl import Workbook #creats writable Workbook
from xlrd import * # Reads .xls
from xlutils.copy import copy
from collections import OrderedDict
import os
from PopulateAmazonTNoS import columnNames, invEnumerated

'''The user is allowed to specify what type of info he is passing.The user should be able to
use these two options:
-Inventory
-Sold
For the inventory option, the user is passing the file will all the inventory.
(I need to make sure all the numbers are correct. This options gives me how many are avaliable).

For the sold the user will pass on a template of the items that were sold.
(I need to add the information to the sold column and calculate the new ammount avaliable).'''



def inventoryDict(sheetInfo, dictionary, items,start):
    alist = [dictionary[items[0]],dictionary[items[1]]]
    adict = OrderedDict()
    cols = sheetInfo.ncols
    rows = sheetInfo.nrows
    for r in range(start,rows):
        currentRowItem = sheetInfo.cell_value(r , alist[0])
        currentRowAmm = sheetInfo.cell_value(r , alist[1])
        adict[currentRowItem] = currentRowAmm
    return adict
            
                
    

if __name__ == '__main__':


    amazonInfoFileName = "AmazonShort.xls"
    amazonInfoType = "Inventory"
    amazonInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", amazonInfoFileName)
    amazonInfoSheet = 1
    aColNameStart = 1
    aInfoStart = 2

    ebayInfoFileName = "EbayShort.xls"
    ebayInfoType = "Inventory"
    ebayInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", ebayInfoFileName)
    ebayInfoSheet = 1
    eColNameStart = 1
    eInfoStart = 2

    websiteInfoFileName = "WebsiteShort.xls"
    websiteInfoType = "Inventory"
    websiteInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", websiteInfoFileName)
    websiteInfoSheet = 1
    wColNameStart = 1
    wInfoStart = 2

    posInfoFile = "POS.xls"
    posInfoType = "Inventory"
    posInfoLocal = os.path.join(os.getcwd(), "ToUpdate", "Shortened", posInfoFile)
    posInfoSheet = 1
    pColNameStart = 1
    pInfoStart = 2
    
    inventoryFileName = "InventoryShort.xls"
    inventoryInfoLocal = os.path.join(os.getcwd(), "MainFiles", "InventoryFile", inventoryFileName)
    invSheetNum = 4
    invColNameStart = 10
    invInfoStart = 11

    
    #Here is where i change the indexes to match
    amazonInfoSheet = amazonInfoSheet - 1
    aColNameStart = aColNameStart - 1
    aInfoStart = aInfoStart - 1

    ebayInfoSheet = ebayInfoSheet - 1
    eColNameStart = eColNameStart - 1
    eInfoStart = eInfoStart - 1
    
    websiteInfoSheet = websiteInfoSheet - 1
    wColNameStart = wColNameStart - 1
    wInfoStart = wInfoStart - 1

    posInfoSheet = posInfoSheet - 1
    pColNameStart = pColNameStart - 1
    pInfoStart = pInfoStart - 1

    invSheetNum = invSheetNum - 1
    invColNameStart = invColNameStart - 1
    invInfoStart = invInfoStart - 1
    #end of index change



    aInfoCols = ["sku","quantity"]
    amazonColNames = columnNames(amazonInfoLocal, aColNameStart, amazonInfoSheet)
    amazonInfoOpen = open_workbook(amazonInfoLocal)
    amazonSheet = amazonInfoOpen.sheet_by_index(amazonInfoSheet)
    amazonInv = inventoryDict(amazonSheet, amazonColNames, aInfoCols, aInfoStart)
    print("Amazon INV")
    print(amazonInv)
    print()


    eInfoCols = ["Custom Label", "Quantity Available"]
    ebayColNames = columnNames(ebayInfoLocal, eColNameStart, ebayInfoSheet)
    ebayInfoOpen = open_workbook(ebayInfoLocal)
    ebaySheet = ebayInfoOpen.sheet_by_index(ebayInfoSheet)
    ebayInv = inventoryDict(ebaySheet, ebayColNames, eInfoCols, eInfoStart)
    print("Ebay INV")
    print(ebayInv)
    print()

    

    webInfoCols = ["Product ID", "Stock"]
    webColNames = columnNames(websiteInfoLocal, wColNameStart, websiteInfoSheet)
    websiteInfoOpen = open_workbook(websiteInfoLocal)
    websiteSheet = websiteInfoOpen.sheet_by_index(websiteInfoSheet)
    webInv = inventoryDict(websiteSheet, webColNames, webInfoCols, wInfoStart)
    print("Website INV")
    print(webInv)
    print()

    posInfoCols = ["Item Name", "Qty 1"]
    posColNames = columnNames(posInfoLocal, pColNameStart,posInfoSheet)
    posInfoOpen = open_workbook(posInfoLocal)
    posSheet = posInfoOpen.sheet_by_index(posInfoSheet)
    posInv = inventoryDict(posSheet, posColNames, posInfoCols, pInfoStart)
    print("POS Inv")
    print(posInv)
    print()
    
    newWorkbook = Workbook()
    wSheet = newWorkbook.active

    with open_workbook(inventoryInfoLocal) as openInventory:
        inventorySheet = openInventory.sheet_by_index(invSheetNum)
        rows = inventorySheet.nrows
        cols = inventorySheet.ncols
        
        columnNames = columnNames(inventoryInfoLocal, invColNameStart, invSheetNum)
        print(columnNames)
        revColNames = invEnumerated(columnNames)
        
        
        for k,v in columnNames.items():
            
            currentCell = wSheet.cell(row = 1, column = v + 1)
            currentCell.value = k
        for r in range(invInfoStart,rows):
            print("NEW PRODUCT")
            startAmount = 0
            avaliablePos = 0
            avaliableWeb = 0
            avaliableAma = 0
            avaliableEbay = 0

            soldPos = 0
            soldAma = 0
            soldWeb = 0
            soldEbay = 0
            totalSold = 0
            finalAmount = 0
            for c in range(cols):
                #print(r,c)
                currentCell = wSheet.cell(row = r - invInfoStart + 2 ,column = c+1)
                if c == columnNames["Item Number"]:
                    #print(inventorySheet.cell_value(r,c))
                    currentCell.value = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["POS Item Name"]:
                    #print(inventorySheet.cell_value(r,c))
                    avaliablePos = posInv[inventorySheet.cell_value(r,c)]
                    currentCell.value = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Amazon Sku"]:
                    avaliableAma = amazonInv[inventorySheet.cell_value(r,c)]
                    currentCell.value = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Website Item Name"]:
                    avaliableWeb = webInv[inventorySheet.cell_value(r,c)]
                    currentCell.value = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Ebay Custom Label"]:
                    avaliableEbay = ebayInv[inventorySheet.cell_value(r,c)]
                    currentCell.value = inventorySheet.cell_value(r,c)
                    
                if c == columnNames["Short Description"]:
                    currentCell.value = inventorySheet.cell_value(r ,c)
                    
                if c == columnNames["Starting Quantity"]:
                    startAmount = inventorySheet.cell_value(r ,c)
                    currentCell.value = startAmount

                if c == columnNames["Sold in Store"]:
                    if inventorySheet.cell_value(r ,c) == 0:
                        soldPos = startAmount - avaliablePos
                        currentCell.value = soldPos
                    
                if c == columnNames["Sold in Amazon"]:
                    if inventorySheet.cell_value(r ,c) == 0:
                        soldAma = startAmount - avaliableAma
                        currentCell.value = soldAma
                    
                if c == columnNames["Sold in Ebay"]:
                    if inventorySheet.cell_value(r ,c) == 0:
                        soldEbay = startAmount - avaliableEbay
                        currentCell.value = soldEbay
                    
                if c == columnNames["Sold in Website"]:
                    if inventorySheet.cell_value(r ,c) == 0:
                        soldWeb = startAmount - avaliableWeb
                        currentCell.value = soldWeb
                
                if c == columnNames["Total Sold"]:
                    if inventorySheet.cell_value(r ,c) == 0:
                        totalSold = soldPos + soldAma + soldEbay + soldWeb
                        currentCell.value = totalSold
                if c == columnNames["Final Amount Avaliable"]:
                    currentCell.value = startAmount - totalSold
                

                    


    newWorkbook.save("Inventory.xls")
    print("Saved!")
            
    

